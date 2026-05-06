# -*- coding: utf-8 -*-
"""Generate Excel schedule from TCM tech plan v2.1 — 2 calendar hours per day."""
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "qiyan_nexus_plan_v2.1_schedule.xlsx"

# WBS: phase, task name, description/deliverable, estimated hours (inferred from doc scope)
ROWS = [
    # --- 前期：对照 V2.1 文档与设计冻结 ---
    ("前期·研读", "V2.1 变更矩阵落地", "对照「零」章：中文 Embedding、认证、PDF、PgBouncer、Flower、R2、G6、合规与变现路径差异清单确认", 6),
    ("前期·研读", "产品定位与边界评审", "「一」章：特应性皮炎垂直定位、不做清单与利益相关方共识纪要", 4),
    ("前期·研读", "技术栈总览冻结", "「二」章：Next.js15/FastAPI/pgvector/Neo4j/Celery 等选型签字版", 4),
    ("前期·研读", "架构与路由策略文档化", "「三」章：整体架构图、LLM 双模型路由、语义缓存伪代码落设计文档", 10),
    ("前期·研读", "遗留代码修复方案验证", "「四」章：pgvector 注册向量、gseapy 富集、Celery 异步接口 PoC", 12),
    ("前期·研读", "双语 Embedding 设计评审", "「五」章：768 维统一表、language 字段、langdetect 分流策略评审", 8),
    ("前期·研读", "TCM 数据与 MVP 数据清单", "「六」章：HERB/SymMap/文献规模（1000+500）与优先级的执行拆解", 10),
    ("前期·研读", "基础设施五模块方案确认", "「七」章：Auth/PDF/R2/G6/监控 免费方案选型会议纪要", 8),
    ("前期·研读", "部署与资源预算对齐", "「八」章：Docker 内存分配、SWAP/升配决策、Nginx 限流策略确认", 8),
    # --- MVP 第1月（文档「九」+「十二」快速启动重叠部分）---
    ("MVP·第1月", "账号与云服务开户", "DeepSeek/Claude/Neo4j Aura/Cloudflare R2/阿里云 学生机（对应十二章 Day0）", 4),
    ("MVP·第1月", "本地开发栈安装", "Python 3.11、Node 20、Docker Desktop、VS Code 与仓库初始化", 4),
    ("MVP·第1月", "FastAPI 工程与依赖安装", "fastapi、uvicorn、pgvector、celery、sentence-transformers 等（十二章 Day1）", 8),
    ("MVP·第1月", "环境变量与 /docs 验证", ".env、API Key、OpenAPI 可访问", 4),
    ("MVP·第1月", "PostgreSQL+pgvector 与 papers 表", "768 维、language 字段、register_vector 写入测试（十二章 Day2）", 12),
    ("MVP·第1月", "PgBouncer 联调", "transaction 模式、后端经池连接压测", 8),
    ("MVP·第1月", "Redis 与向量存取回归", "缓存键空间、向量插入查询 SLI", 6),
    ("MVP·第1月", "双模型权重下载与缓存", "text2vec-chinese + PubMedBERT 首次拉取与冷启动策略", 10),
    ("MVP·第1月", "文献样本与 PyMuPDF 管道", "≥50 篇 PDF、扫描件降级 unstructured（十二章 Day3）", 14),
    ("MVP·第1月", "批量导入与 /rag/answer", "中英分流 Embedding、检索问答端到端", 16),
    ("MVP·第1月", "Celery+gseapy 药理学异步骨架", "analyze/result、Flower 5555、Nginx 限流草案（十二章 Day4）", 18),
    ("MVP·第1月", "Next.js 壳与 Ant Design 引入", "创建前端工程、基础布局（十二章 Day5 部分）", 10),
    ("MVP·第1月", "联调与首批部署", "Compose 含 PgBouncer/Flower、阿里云+Nginx+SSL、3–5 人抽测（十二章 Day6-7）", 20),
    ("MVP·第1月", "里程碑验收：基础 RAG", "API 文档可访问，中英文论文问答可用（文档「九」第1月）", 6),
    # --- MVP 第2月 ---
    ("MVP·第2月", "NextAuth 邮件魔法链接", "邀请白名单、登录回调与权限中间件", 16),
    ("MVP·第2月", "前端文献检索与任务轮询", "task_id 轮询、Ant Design 表单与状态展示", 18),
    ("MVP·第2月", "AntV G6 技术预研与 Demo", "网络图最小示例、与后端 mock 数据联调", 14),
    ("MVP·第2月", "PDF 上传至 R2 全链路", "预签名/服务端上传、元数据入库", 16),
    ("MVP·第2月", "解析入库与知识库更新", "上传后触发解析、分块、Embedding 流水线", 18),
    ("MVP·第2月", "里程碑验收：登录+上传+检索", "邀请登录、上传论文、检索回答（「九」第2月）", 8),
    # --- MVP 第3月 ---
    ("MVP·第3月", "网络药理学编排深化", "方剂输入、成分-靶点-通路业务规则与错误处理", 20),
    ("MVP·第3月", "Celery 队列监控与重试", "任务超时、失败队列、Flower 运维手册", 12),
    ("MVP·第3月", "Cloudflare R2 生产配置", "生命周期、CORS、与后端密钥轮换", 10),
    ("MVP·第3月", "异步报告生成与 Claude 配额", "Redis 每用户每日限额、语义缓存接入（文档 3.3）", 18),
    ("MVP·第3月", "里程碑验收：异步药理学+R2", "输入方剂异步分析、PDF 托管稳定（「九」第3月）", 8),
    # --- MVP 第4月 ---
    ("MVP·第4月", "Neo4j Aura 建模与导入", "中药-成分-靶点、与 HERB/TCM-KG 对齐", 22),
    ("MVP·第4月", "图谱查询 API", "Cypher 服务层、分页与权限", 14),
    ("MVP·第4月", "Neo4j Bloom 演示环境", "内测演示用 Bloom 视图与口径说明", 8),
    ("MVP·第4月", "AntV G6 正式可视化页", "替代/并行 Bloom 的产品化网络探索 UI", 22),
    ("MVP·第4月", "里程碑验收：图谱可查可展示", "「九」第4月：图谱查询+图形化展示", 6),
    # --- MVP 第5月 ---
    ("MVP·第5月", "内测招募与邀请运营", "50 位用户池、反馈渠道、问卷与访谈提纲", 12),
    ("MVP·第5月", "Sentry 全栈接入", "Python+Next 异常、版本区分环境", 10),
    ("MVP·第5月", "BetterStack 可用性监控", "关键接口探测、告警渠道", 8),
    ("MVP·第5月", "稳定性与容量观测", "Flower+日志聚合、慢查询与 Redis 监控项", 14),
    ("MVP·第5月", "反馈闭环与迭代清单", "≥100 条反馈整理、优先级排序", 12),
    ("MVP·第5月", "里程碑验收：内测与监控", "有效用户≥50、零宕机目标尽力（「九」第5月）", 8),
    # --- MVP 第6月 ---
    ("MVP·第6月", "缺陷收敛与回归测试", "P0/P1 缺陷清零冲刺", 18),
    ("MVP·第6月", "付费报告导出（PDF）", "微信支付/支付宝对接、订单与开票信息最小集", 24),
    ("MVP·第6月", "报告模板与学术规范", "成分-靶点-通路+参考文献+图表（「九」变现细节）", 16),
    ("MVP·第6月", "收费策略与运营物料", "¥99–199 定价页、使用条款与免责声明展示", 10),
    ("MVP·第6月", "里程碑验收：首笔收入", "第一笔付费科研报告入账目标（「九」第6月）", 8),
    # --- 资金、合规与持续运营 ---
    ("持续·合规预算", "MVP 预算执行与账单巡检", "「十」章：DeepSeek/Claude/服务器费用台账与告警阈值", 8),
    ("持续·合规预算", "风险矩阵与合规落地", "「十一」章：PIPL、免责声明、诊疗边界文案与算法备案路线图", 14),
    ("持续·合规预算", "文档与运维交付件归档", "架构图、Runbook、备份与灾备说明 v1", 10),
]


def main():
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="D9E2F3")

    wb = Workbook()
    ws = wb.active
    ws.title = "排期总表"

    headers = [
        "序号",
        "阶段/模块",
        "任务名称",
        "描述/交付物",
        "预估工时(h)",
        "占用日历天数",
        "累计天数(至本任务结束)",
        "开始日序号",
        "结束日序号",
        "备注",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    start_day = 1
    total_hours = 0.0
    for i, (phase, name, desc, hours) in enumerate(ROWS, 1):
        total_hours += hours
        days = max(1, int(math.ceil(hours / 2.0)))
        end_day = start_day + days - 1

        row_vals = [
            i,
            phase,
            name,
            desc,
            round(hours, 1),
            days,
            end_day,
            start_day,
            end_day,
            "每日有效工时按 2h；日序号从第 1 天起算，任务顺序衔接",
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (1, 5, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # zebra by phase
        if i % 2:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = sub_fill

        start_day = end_day + 1

    # Summary sheet
    ws2 = wb.create_sheet("说明与假设", 1)
    ws2["A1"] = "生成说明"
    ws2["A1"].font = Font(bold=True, size=12)
    notes = [
        "来源文档：qiyan_nexus_plan_v2.1.docx（Qiyan Nexus技术架构方案 V2.1）。",
        "工时：文档未给出详细工时的部分，按章节范围与「九、产品路线图」「十二、第一周快速启动清单」推断。",
        "排期规则：总日历天数 = ceil(预估工时 ÷ 2)；同一任务连续占用从「开始日序号」到「结束日序号」。",
        "「开始日序号」第 1 天 = 项目启动第 1 个日历日；未扣节假日（可按需后续压缩/拉长）。",
        f"本表任务数：{len(ROWS)}；预估总工时：{total_hours:.1f} h；按 2h/日约合 {math.ceil(total_hours/2)} 个日历日。",
    ]
    for r, t in enumerate(notes, 2):
        ws2.cell(row=r, column=1, value=t)
        ws2.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 96

    # column widths
    widths = [6, 14, 22, 52, 12, 14, 22, 12, 12, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
